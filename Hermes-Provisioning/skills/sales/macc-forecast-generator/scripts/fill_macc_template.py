"""CLI entry point: the one real, mechanical write path for a customer's
own MACC Forecast Sheet (2026-08-24, rewritten to match the real
hand-built template -- see `Work/Templates/MACC Forecast Template.md`
for the full cell/table contract this script keys off of; read it before
changing this script, the two must stay in sync by hand).

This script does NOT compute anything -- it copies the real template and
sets real input cells/table rows on `Config` and `Forecast` only. Every
quarterly total, both growth scenarios, both dashboards, and the chart
are real Excel formulas already in the template -- they recalculate the
moment the file is opened.

**Real design point (2026-08-24): an engagement's quarterly consumption
is entered DIRECTLY per quarter, not derived from a start date and one
flat monthly figure.** Gather (or estimate, per real knowledge of the
deal) a real number for each active quarter -- ramping, flat, seasonal,
whatever the real shape actually is -- never mechanically project one
number forward.

Usage:
    python fill_macc_template.py --vault-path P --input-file F

F: {
  "customer": str,
  "term_months": int,                    # e.g. 60 for a 5-year deal
  "forecast_start_quarter": str,         # "YYYY-MM-01" -- first month of Q1
  "last_macc_size_usd": float,
  "last_macc_consumption_end_usd": float,  # quarterly run-rate
  "last_macc_end_date": str,             # "YYYY-MM-DD"
  "assumed_qoq_growth_rate": float,      # e.g. 0.03 for 3%
  "growth_history": [                     # optional, real past quarters, oldest first
    {"quarter": "2025-Q3", "actual_consumption_usd": 38000},
    ...
  ],                                       # max 12
  "engagements": [
    {
      "name": str,
      "type": str,
      "source": str | null,               # e.g. "Opportunity: Azure Data Manager for Energy" or "asked directly"
      "quarterly_consumption_usd": [float, ...]  # one value per active quarter, Q1 first
    },
    ...
  ]                                        # max 8 without extending the table -- see _ensure_table_capacity
}

Prints {"created": true, "path": str} or {"error": str}.

IMPORTANT -- run with PYTHONUTF8=1 set: this Python install silently
corrupts a non-ASCII character (an accented name, etc.) written into an
xlsx otherwise -- confirmed live, 2026-08-24, same class of gremlin
found repeatedly elsewhere in this codebase.
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

_SLUG_INVALID_CHARS = re.compile(r'[\\/:*?"<>|]')
_TEMPLATE_RELATIVE_PATH = ("Work", "Templates", "MACC Forecast Template.xlsx")

_MAX_QUARTERS = 20  # template's own fixed capacity (Forecast!D:W)
_MAX_GROWTH_HISTORY_ROWS = 12
# The Engagements table's own real starter rows (Forecast!6:13) sit
# directly above the "Quarterly Forecast" section starting at row 16 --
# NO spare rows exist between them. A hard cap, not an
# auto-expand-and-hope: found live, 2026-08-24, extending the table past
# 8 rows silently grew it INTO that section (`Table.ref` extended past
# row 13 with no collision check), corrupting the sheet with no error.
_MAX_ENGAGEMENTS = 8

# Real, fixed cell/table addresses -- must match `build_template.md`'s
# own documented contract (and the hand-built workbook itself) exactly.
CONFIG_CUSTOMER = "B5"
CONFIG_TERM_MONTHS = "B6"
CONFIG_FORECAST_START = "B9"
CONFIG_LAST_MACC_SIZE = "B13"
CONFIG_LAST_MACC_CONSUMPTION_END = "B14"
CONFIG_LAST_MACC_END_DATE = "B15"
CONFIG_ASSUMED_QOQ_GROWTH = "B16"
GROWTH_HISTORY_TABLE_NAME = "QoQGrowthHistory"
GROWTH_HISTORY_FIRST_DATA_ROW = 20  # header is row 19

ENGAGEMENTS_TABLE_NAME = "Engagements"
ENGAGEMENTS_HEADER_ROW = 5
ENGAGEMENTS_FIRST_DATA_ROW = 6
ENGAGEMENTS_FIRST_QUARTER_COL = 4  # column D = Q1


def _slugify(text: str, max_len: int = 80) -> str:
    slug = _SLUG_INVALID_CHARS.sub("-", text).strip()
    return slug[:max_len] if slug else "Untitled"


def _ensure_table_capacity(ws, table_name: str, needed_data_rows: int, header_row: int) -> None:
    """Extends a real Excel Table's own `ref` range to cover
    `needed_data_rows` if the template's starter size is smaller --
    openpyxl doesn't auto-grow a Table when you write values below its
    current range, unlike Excel's own UI when a person types into the
    row right after a Table."""
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    current_data_rows = max_row - header_row
    if needed_data_rows <= current_data_rows:
        return
    new_max_row = header_row + needed_data_rows
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"


def fill_macc_template(
    vault_path: Path,
    output_path: Path,
    customer: str,
    term_months: int,
    forecast_start_quarter: str,
    last_macc_size_usd: float,
    last_macc_consumption_end_usd: float,
    last_macc_end_date: str,
    assumed_qoq_growth_rate: float,
    growth_history: list[dict],
    engagements: list[dict],
) -> dict:
    if not customer.strip():
        return {"error": "customer is required"}
    if term_months <= 0:
        return {"error": "term_months must be positive"}
    if not engagements:
        return {"error": "at least one engagement is required"}
    if len(engagements) > _MAX_ENGAGEMENTS:
        return {"error": f"at most {_MAX_ENGAGEMENTS} engagements fit this template (got {len(engagements)}) -- the Quarterly Forecast section sits directly below with no spare rows, say so plainly rather than trying to cram more in"}
    if len(growth_history) > _MAX_GROWTH_HISTORY_ROWS:
        return {"error": f"at most {_MAX_GROWTH_HISTORY_ROWS} growth history rows fit this template (got {len(growth_history)})"}
    for eng in engagements:
        if len(eng.get("quarterly_consumption_usd") or []) > _MAX_QUARTERS:
            return {"error": f"at most {_MAX_QUARTERS} quarters fit this template (got {len(eng['quarterly_consumption_usd'])} for {eng.get('name')!r})"}

    template_path = vault_path.joinpath(*_TEMPLATE_RELATIVE_PATH)
    if not template_path.is_file():
        return {"error": f"template not found at {template_path}"}

    wb = load_workbook(template_path)

    # ── Config ────────────────────────────────────────────────────────
    cfg = wb["Config"]
    cfg[CONFIG_CUSTOMER] = customer
    cfg[CONFIG_TERM_MONTHS] = term_months
    try:
        cfg[CONFIG_FORECAST_START] = datetime.strptime(forecast_start_quarter[:10], "%Y-%m-%d").date()
    except ValueError:
        return {"error": f"forecast_start_quarter must be YYYY-MM-DD, got {forecast_start_quarter!r}"}
    cfg[CONFIG_LAST_MACC_SIZE] = float(last_macc_size_usd)
    cfg[CONFIG_LAST_MACC_CONSUMPTION_END] = float(last_macc_consumption_end_usd)
    try:
        cfg[CONFIG_LAST_MACC_END_DATE] = datetime.strptime(last_macc_end_date[:10], "%Y-%m-%d").date()
    except ValueError:
        return {"error": f"last_macc_end_date must be YYYY-MM-DD, got {last_macc_end_date!r}"}
    cfg[CONFIG_ASSUMED_QOQ_GROWTH] = float(assumed_qoq_growth_rate)

    if growth_history:
        _ensure_table_capacity(cfg, GROWTH_HISTORY_TABLE_NAME, len(growth_history), GROWTH_HISTORY_FIRST_DATA_ROW - 1)
        for i, entry in enumerate(growth_history):
            r = GROWTH_HISTORY_FIRST_DATA_ROW + i
            cfg.cell(r, 1, entry.get("quarter") or "")
            cfg.cell(r, 2, float(entry.get("actual_consumption_usd") or 0))
            # Column C (QoQ Growth %) is a formula already baked into the
            # template for every row up to its own real capacity -- never
            # overwritten here.

    # ── Forecast: Engagements table ──────────────────────────────────
    # No _ensure_table_capacity call here, deliberately -- the
    # _MAX_ENGAGEMENTS guard above already caps real engagement count at
    # the table's own true safe capacity (8 rows, nothing below it to
    # collide with); extending this specific table is never safe (see
    # _MAX_ENGAGEMENTS's own comment).
    fc = wb["Forecast"]
    for i, eng in enumerate(engagements):
        r = ENGAGEMENTS_FIRST_DATA_ROW + i
        fc.cell(r, 1, eng.get("name") or "Untitled Engagement")
        fc.cell(r, 2, eng.get("type") or "")
        fc.cell(r, 3, eng.get("source") or "")
        quarterly = eng.get("quarterly_consumption_usd") or []
        for q, value in enumerate(quarterly):
            col = ENGAGEMENTS_FIRST_QUARTER_COL + q
            cell = fc.cell(r, col, float(value))
            cell.number_format = '"$"#,##0'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {"created": True, "path": str(output_path)}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--vault-path", required=True)
    parser.add_argument("--input-file", required=True)
    args = parser.parse_args()

    vault_path = Path(args.vault_path)
    data = json.loads(Path(args.input_file).read_text(encoding="utf-8-sig"))

    customer = data.get("customer", "")
    generated_date = datetime.now().strftime("%Y-%m-%d")
    customer_dir = vault_path / "Work" / "Customers" / _slugify(customer) / "Files" / "MACC Estimator"
    output_path = customer_dir / f"{_slugify(customer)} MACC Forecast {generated_date}.xlsx"

    result = fill_macc_template(
        vault_path,
        output_path,
        customer=customer,
        term_months=int(data.get("term_months") or 60),
        forecast_start_quarter=data.get("forecast_start_quarter") or generated_date[:8] + "01",
        last_macc_size_usd=data.get("last_macc_size_usd") or 0,
        last_macc_consumption_end_usd=data.get("last_macc_consumption_end_usd") or 0,
        last_macc_end_date=data.get("last_macc_end_date") or generated_date,
        assumed_qoq_growth_rate=data.get("assumed_qoq_growth_rate") or 0,
        growth_history=data.get("growth_history") or [],
        engagements=data.get("engagements") or [],
    )
    print(json.dumps(result, ensure_ascii=False))
    return 0 if "error" not in result else 1


if __name__ == "__main__":
    raise SystemExit(main())
