"""ONE-TIME builder for the real MACC Forecast Excel TEMPLATE (2026-08-24).
This script is NOT the ongoing per-customer Skill -- run it once to
construct the real workbook; the ongoing Skill (`fill_macc_template.py`,
same folder) only ever copies this file and sets real input cells/table
rows. Every total, projection, and chart is a REAL EXCEL FORMULA -- this
script computes nothing itself.

Real design history (operator, 2026-08-24, in order):
1. "Instead of Creating a Python File for that as a Skill We should
   Create an Excel with Calculation and the MACC Agent fill that sheet"
   -- Excel formulas do the math, not Python.
2. "The Table Should be Baseline at the bottom, Engagements above it
   Quarterly... I can Add More Engagements Format that as a Table then
   Have a Sheet with the Results then a Sheet with Charts" -- a real,
   growable Excel Table for Engagements (not a fixed row count), a
   quarterly (not monthly) Baseline projection below it, Results and
   Charts as their own sheets.
3. "We can Have a Sheet for Extra Configurable Parameters as well" ->
   "For example history of QoQ Growth, Last Macc Size Last MACC
   Consumption End QoQ Growth Will Affect How the Macc Size will go" --
   the Baseline isn't just bottom-up engagements, it's an EXISTING
   BUSINESS TRAJECTORY (seeded from the customer's own last MACC,
   compounding at an assumed QoQ growth rate) PLUS new engagements
   layered on top once each one's own start date arrives.

Four real sheets:
  - "Inputs"       -- header info + the real, growable Engagements Table
                       + the quarterly Baseline projection grid.
  - "Config"        -- Last MACC Size/Consumption End/End Date, a real
                       QoQ Growth History table, and one explicit
                       Assumed Forward QoQ Growth % that drives the
                       trajectory (never silently averaged from history).
  - "Results"       -- the real computed summary (Total MACC Size, the
                       existing-trajectory vs. new-engagement split).
  - "Charts"        -- the chart, on its own sheet.

Capacity (fixed -- matches `Work/Templates/MACC Forecast Template.md`'s
own documented contract): quarters up to 20 (5 years, the real common
MACC term ceiling); Engagements is a real Excel Table, so it grows in
Excel itself, not a fixed row count here (started with 8 blank rows).

Usage:
    python build_macc_template.py --output-path P
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

_ENGAGEMENT_START_ROWS = 8  # real Table, grows in Excel -- this is just the starter size
_QUARTER_SLOTS = 20  # 5 years
_GROWTH_HISTORY_ROWS = 12  # 3 years of quarterly history, real Table
EM_DASH = chr(0x2014)

_NAVY = "1F3864"
_LIGHT_BLUE = "D9E2F3"
_WHITE = "FFFFFF"
_GRAY_BAND = "F2F2F2"
_GREEN = "2E7D32"
_INPUT_YELLOW = "FFF2CC"

_HEADER_FONT = Font(name="Calibri", size=12, bold=True, color=_WHITE)
_TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=_NAVY)
_SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="595959")
_TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color=_NAVY)
_BIG_TOTAL_FONT = Font(name="Calibri", size=16, bold=True, color=_GREEN)

_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_INFO_FILL = PatternFill("solid", fgColor=_LIGHT_BLUE)
_INPUT_FILL = PatternFill("solid", fgColor=_INPUT_YELLOW)

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Real, fixed cell addresses -- the contract `fill_macc_template.py` (and
# the template's own sibling .md) both key off.
CUSTOMER_CELL = "Inputs!B4"
TERM_YEARS_CELL = "Inputs!B5"
TERM_QUARTERS_CELL = "Inputs!B6"
FORECAST_START_CELL = "Inputs!B7"
GENERATED_DATE_CELL = "Inputs!B8"
ENGAGEMENTS_TABLE_NAME = "Engagements"
ENGAGEMENTS_HEADER_ROW = 11
ENGAGEMENTS_FIRST_DATA_ROW = 12
BASELINE_QUARTER_INDEX_ROW = 24  # raw 1..20, used by formulas
BASELINE_QUARTER_LABEL_ROW = 25  # human label, e.g. "2026-Q4"
BASELINE_EXISTING_ROW = 26
BASELINE_NEW_ROW = 27
BASELINE_TOTAL_ROW = 28
BASELINE_FIRST_COL = 2  # column B

LAST_MACC_SIZE_CELL = "Config!B4"
LAST_MACC_CONSUMPTION_END_CELL = "Config!B5"
LAST_MACC_END_DATE_CELL = "Config!B6"
ASSUMED_QOQ_GROWTH_CELL = "Config!B7"
GROWTH_HISTORY_TABLE_NAME = "QoQGrowthHistory"
GROWTH_HISTORY_HEADER_ROW = 10
GROWTH_HISTORY_FIRST_DATA_ROW = 11


def _styled_header(ws, row: int, col: int, text: str):
    cell = ws.cell(row, col, text)
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _build_inputs_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Inputs"

    ws.merge_cells("A1:G1")
    ws["A1"] = f'="MACC Forecast {EM_DASH} "&B4'
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Basic MACC sizing {EM_DASH} fill the yellow input cells and the Engagements table; every projection and chart recalculates automatically."
    ws["A2"].font = _SUBTITLE_FONT

    ws["A4"] = "Customer"
    ws["A5"] = "MACC Term (Years)"
    ws["A6"] = "Term (Quarters)"
    ws["A7"] = "Forecast Start Quarter (first month of Q1)"
    ws["A8"] = "Generated Date"
    for addr in ("A4", "A5", "A6", "A7", "A8"):
        ws[addr].font = _LABEL_FONT
        ws[addr].fill = _INFO_FILL
    ws["B6"] = "=B5*4"
    ws["B6"].font = _TOTAL_FONT
    for addr in ("B4", "B5", "B7", "B8"):
        ws[addr].fill = _INPUT_FILL
        ws[addr].border = _BORDER
    ws["B7"].number_format = "YYYY-MM"
    ws["B8"].number_format = "YYYY-MM-DD"

    # ── Engagements -- a REAL Excel Table, grows when rows are added in
    # Excel itself (operator: "I can Add More Engagements Format that as
    # a Table"). "Start Quarter" is a calculated Table column -- Excel
    # auto-fills it for any new row the same way it auto-fills a Table's
    # own formula columns generally.
    ws.cell(ENGAGEMENTS_HEADER_ROW - 1, 1, "Engagements").font = _LABEL_FONT
    headers = ["Engagement", "Type", "Start Date", "Monthly Consumption (USD)", "Source", "Start Quarter"]
    for c, h in enumerate(headers, start=1):
        ws.cell(ENGAGEMENTS_HEADER_ROW, c, h)
    last_row = ENGAGEMENTS_FIRST_DATA_ROW + _ENGAGEMENT_START_ROWS - 1
    for i in range(_ENGAGEMENT_START_ROWS):
        r = ENGAGEMENTS_FIRST_DATA_ROW + i
        ws.cell(r, 3).number_format = "YYYY-MM-DD"
        ws.cell(r, 4).number_format = '"$"#,##0'
        # Start Quarter (1-based, relative to Inputs!B7) -- blank Start
        # Date -> blank; a Start Date at/before the forecast start
        # counts as Quarter 1.
        ws.cell(r, 6, f'=IF(C{r}="","",IF(C{r}<=$B$7,1,ROUNDDOWN(DATEDIF($B$7,C{r},"m")/3,0)+1))')
    table_ref = f"A{ENGAGEMENTS_HEADER_ROW}:F{last_row}"
    table = Table(displayName=ENGAGEMENTS_TABLE_NAME, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    # ── Baseline (Quarterly) -- BELOW the Engagements table (operator:
    # "The Table Should be Baseline at the bottom Engagements above it
    # Quarterly"). Existing-business trajectory (seeded from Config's own
    # Last MACC figures, compounding at the Assumed QoQ Growth Rate) +
    # new-engagement contributions (summed off the Engagements Table via
    # structured references, so it stays correct as rows are added),
    # summed into one Total row. Quarters beyond the real term (Inputs!B6)
    # show 0, not garbage, regardless of the fixed 20-quarter capacity.
    ws.cell(BASELINE_QUARTER_INDEX_ROW - 1, 1, "Baseline (Quarterly)").font = _LABEL_FONT
    ws.cell(BASELINE_QUARTER_INDEX_ROW, 1, "Quarter #")
    ws.cell(BASELINE_QUARTER_LABEL_ROW, 1, "Quarter")
    ws.cell(BASELINE_EXISTING_ROW, 1, "Existing Business Trajectory")
    ws.cell(BASELINE_NEW_ROW, 1, "New Engagements")
    ws.cell(BASELINE_TOTAL_ROW, 1, "Total")
    ws.cell(BASELINE_TOTAL_ROW, 1).font = _TOTAL_FONT
    for r in (BASELINE_QUARTER_INDEX_ROW, BASELINE_QUARTER_LABEL_ROW, BASELINE_EXISTING_ROW, BASELINE_NEW_ROW, BASELINE_TOTAL_ROW):
        ws.cell(r, 1).border = _BORDER

    for q in range(_QUARTER_SLOTS):
        col = BASELINE_FIRST_COL + q
        col_letter = get_column_letter(col)
        n = q + 1  # 1-based quarter number

        idx_cell = ws.cell(BASELINE_QUARTER_INDEX_ROW, col, n)
        idx_cell.border = _BORDER
        idx_cell.font = _HEADER_FONT
        idx_cell.fill = _HEADER_FILL
        idx_cell.alignment = Alignment(horizontal="center")

        # Human label: forecast-start quarter, chained via EDATE(3 months per step).
        if q == 0:
            label_formula = '=TEXT($B$7,"YYYY")&"-Q"&ROUNDUP(MONTH($B$7)/3,0)'
        else:
            prev_letter = get_column_letter(col - 1)
            label_formula = f'=TEXT(EDATE($B$7,{q}*3),"YYYY")&"-Q"&ROUNDUP(MONTH(EDATE($B$7,{q}*3))/3,0)'
        label_cell = ws.cell(BASELINE_QUARTER_LABEL_ROW, col, label_formula)
        label_cell.border = _BORDER
        label_cell.alignment = Alignment(horizontal="center")

        # Existing trajectory: 0 once beyond the real term; otherwise the
        # seed compounding at the assumed QoQ growth rate for (n-1) quarters.
        existing_formula = (
            f'=IF({col_letter}${BASELINE_QUARTER_INDEX_ROW}>$B$6,0,'
            f'Config!$B$5*(1+Config!$B$7)^({col_letter}${BASELINE_QUARTER_INDEX_ROW}-1))'
        )
        existing_cell = ws.cell(BASELINE_EXISTING_ROW, col, existing_formula)
        existing_cell.number_format = '"$"#,##0'
        existing_cell.border = _BORDER

        # New engagements: 3 months of every engagement whose own Start
        # Quarter has arrived by quarter n, guarded to the real term.
        new_formula = (
            f'=IF({col_letter}${BASELINE_QUARTER_INDEX_ROW}>$B$6,0,'
            f'IFERROR(SUMIF({ENGAGEMENTS_TABLE_NAME}[Start Quarter],"<="&{col_letter}${BASELINE_QUARTER_INDEX_ROW},'
            f'{ENGAGEMENTS_TABLE_NAME}[Monthly Consumption (USD)])*3,0))'
        )
        new_cell = ws.cell(BASELINE_NEW_ROW, col, new_formula)
        new_cell.number_format = '"$"#,##0'
        new_cell.border = _BORDER

        total_cell = ws.cell(BASELINE_TOTAL_ROW, col, f"={col_letter}{BASELINE_EXISTING_ROW}+{col_letter}{BASELINE_NEW_ROW}")
        total_cell.number_format = '"$"#,##0'
        total_cell.font = _TOTAL_FONT
        total_cell.border = _BORDER
        ws.column_dimensions[col_letter].width = 12

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = ws.cell(BASELINE_QUARTER_INDEX_ROW + 1, BASELINE_FIRST_COL)
    return ws


def _build_config_sheet(wb: Workbook):
    ws = wb.create_sheet("Config")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Configurable Parameters"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:D2")
    ws["A2"] = "The existing-business trajectory on the Inputs sheet is seeded and grown from these -- never silently guessed."
    ws["A2"].font = _SUBTITLE_FONT

    labels = [
        ("A4", "Last MACC Size (USD)"),
        ("A5", "Last MACC Consumption End (USD/quarter run-rate)"),
        ("A6", "Last MACC End Date"),
        ("A7", "Assumed Forward QoQ Growth Rate (%)"),
    ]
    for addr, label in labels:
        ws[addr] = label
        ws[addr].font = _LABEL_FONT
        ws[addr].fill = _INFO_FILL
    for addr in ("B4", "B5", "B6", "B7"):
        ws[addr].fill = _INPUT_FILL
        ws[addr].border = _BORDER
    ws["B4"].number_format = '"$"#,##0'
    ws["B5"].number_format = '"$"#,##0'
    ws["B6"].number_format = "YYYY-MM-DD"
    ws["B7"].number_format = "0.0%"

    # ── QoQ Growth History -- a real Table, Mahmoud/macc-expert fills in
    # real past quarters' actual consumption; QoQ Growth % is a
    # calculated Table column (auto-fills for new rows), never an input.
    ws.cell(GROWTH_HISTORY_HEADER_ROW - 1, 1, "QoQ Growth History (real past quarters, oldest first)").font = _LABEL_FONT
    headers = ["Quarter", "Actual Consumption (USD)", "QoQ Growth %"]
    for c, h in enumerate(headers, start=1):
        ws.cell(GROWTH_HISTORY_HEADER_ROW, c, h)
    last_row = GROWTH_HISTORY_FIRST_DATA_ROW + _GROWTH_HISTORY_ROWS - 1
    for i in range(_GROWTH_HISTORY_ROWS):
        r = GROWTH_HISTORY_FIRST_DATA_ROW + i
        ws.cell(r, 2).number_format = '"$"#,##0'
        ws.cell(r, 3).number_format = "0.0%"
        if i == 0:
            ws.cell(r, 3, "")  # no prior row to compare the very first history row against
        else:
            prev_r = r - 1
            ws.cell(r, 3, f'=IF(OR(B{r}="",B{prev_r}="",B{prev_r}=0),"",(B{r}-B{prev_r})/B{prev_r})')
    growth_table = Table(displayName=GROWTH_HISTORY_TABLE_NAME, ref=f"A{GROWTH_HISTORY_HEADER_ROW}:C{last_row}")
    growth_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(growth_table)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    return ws


def _build_results_sheet(wb: Workbook):
    ws = wb.create_sheet("Results")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Results"
    ws["A1"].font = _TITLE_FONT

    value_font = Font(name="Calibri", size=11)
    last_col_letter = get_column_letter(BASELINE_FIRST_COL + _QUARTER_SLOTS - 1)
    first_col_letter = get_column_letter(BASELINE_FIRST_COL)
    rows = [
        ("Estimated Total MACC Size", f"=SUM(Inputs!{first_col_letter}{BASELINE_TOTAL_ROW}:{last_col_letter}{BASELINE_TOTAL_ROW})", _BIG_TOTAL_FONT, '"$"#,##0'),
        ("  From Existing Business Trajectory", f"=SUM(Inputs!{first_col_letter}{BASELINE_EXISTING_ROW}:{last_col_letter}{BASELINE_EXISTING_ROW})", value_font, '"$"#,##0'),
        ("  From New Engagements", f"=SUM(Inputs!{first_col_letter}{BASELINE_NEW_ROW}:{last_col_letter}{BASELINE_NEW_ROW})", value_font, '"$"#,##0'),
        ("Customer", "=Inputs!B4", value_font, None),
        ("MACC Term", '=Inputs!B5&" years ("&Inputs!B6&" quarters)"', value_font, None),
        ("Last MACC Size", "=Config!B4", value_font, '"$"#,##0'),
        ("Assumed Forward QoQ Growth", "=Config!B7", value_font, "0.0%"),
    ]
    for i, (label, formula, font, numfmt) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, label).font = _LABEL_FONT
        cell = ws.cell(r, 2, formula)
        cell.font = font
        if numfmt:
            cell.number_format = numfmt
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 26
    return ws


def _build_charts_sheet(wb: Workbook, inputs_ws):
    ws = wb.create_sheet("Charts")
    chart = LineChart()
    chart.title = f"Quarterly Consumption {EM_DASH} Existing Trajectory vs. New Engagements vs. Total"
    chart.style = 2
    chart.y_axis.title = "USD / quarter"
    chart.x_axis.title = "Quarter"

    last_col = BASELINE_FIRST_COL + _QUARTER_SLOTS - 1
    cats_ref = Reference(inputs_ws, min_col=BASELINE_FIRST_COL, max_col=last_col, min_row=BASELINE_QUARTER_LABEL_ROW, max_row=BASELINE_QUARTER_LABEL_ROW)
    for row in (BASELINE_EXISTING_ROW, BASELINE_NEW_ROW, BASELINE_TOTAL_ROW):
        data_ref = Reference(inputs_ws, min_col=BASELINE_FIRST_COL, max_col=last_col, min_row=row, max_row=row)
        chart.add_data(data_ref, titles_from_data=False, from_rows=True)
    chart.set_categories(cats_ref)
    chart.series[0].tx = SeriesLabel(v="Existing Business Trajectory")
    chart.series[1].tx = SeriesLabel(v="New Engagements")
    chart.series[2].tx = SeriesLabel(v="Total")
    chart.width = 30
    chart.height = 14
    ws.add_chart(chart, "A2")
    return ws


def build_template(output_path: Path) -> None:
    wb = Workbook()
    inputs_ws = _build_inputs_sheet(wb)
    _build_config_sheet(wb)
    _build_results_sheet(wb)
    _build_charts_sheet(wb, inputs_ws)
    wb.active = 0  # land on Inputs when opened

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-path", required=True)
    args = parser.parse_args()
    build_template(Path(args.output_path))
    print(f"Built template at {args.output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
